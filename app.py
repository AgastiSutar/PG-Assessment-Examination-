import os

from flask import Flask, render_template, request
from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook

from questions import questions

from config import (
    EXAM_DURATION,
    MARKS_PER_QUESTION,
    NEGATIVE_MARKING,
    PASS_PERCENTAGE
)


app = Flask(__name__)


# ==========================
# HOME PAGE
# ==========================

@app.route("/")
def exam():

    return render_template("exam.html")


# ==========================
# START EXAMINATION
# ==========================

@app.route("/start_exam", methods=["POST"])
def start_exam():

    name = request.form["name"]
    roll = request.form["roll"]

    return render_template(
        "test.html",
        questions=questions,
        student_name=name,
        roll_number=roll,
        duration=EXAM_DURATION
    )


# ==========================
# SAVE RESULT TO EXCEL FILE
# ==========================

def save_result(name, roll, marks):

    file_name = "PG Assessment Result.xlsx"

    if os.path.exists(file_name):

        workbook = load_workbook(file_name)
        sheet = workbook.active

    else:

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"

        sheet.append(
            ["Name", "Roll Number", "Marks"]
        )

    sheet.append(
        [name, roll, marks]
    )

    workbook.save(file_name)


# ==========================
# SUBMIT EXAMINATION
# ==========================

@app.route("/submit_exam", methods=["POST"])
def submit_exam():

    total_questions = len(questions)

    attempted = 0
    correct = 0
    wrong = 0
    unanswered = 0

    marks = 0

    for index, question in enumerate(questions, start=1):

        selected = request.form.get(
            f"question_{index}"
        )

        if selected is None:

            unanswered += 1
            continue

        attempted += 1

        if selected == question["answer"]:

            correct += 1
            marks += MARKS_PER_QUESTION

        else:

            wrong += 1
            marks -= NEGATIVE_MARKING

    total_marks = (
        total_questions *
        MARKS_PER_QUESTION
    )

    percentage = (
        (marks / total_marks) * 100
    )

    if percentage >= PASS_PERCENTAGE:

        status = "PASS"

    else:

        status = "FAIL"

    student_name = request.form.get(
        "student_name"
    )

    roll_number = request.form.get(
        "roll_number"
    )

    save_result(

        student_name,

        roll_number,

        round(marks, 2)

    )

    current_date = datetime.now().strftime(
        "%d-%m-%Y"
    )

    current_time = datetime.now().strftime(
        "%I:%M %p"
    )

    return render_template(

        "result.html",

        student_name=student_name,

        roll_number=roll_number,

        date=current_date,

        time=current_time,

        attempted=attempted,

        correct=correct,

        wrong=wrong,

        unanswered=unanswered,

        marks=round(marks, 2),

        total_marks=round(
            total_marks, 2
        ),

        percentage=round(
            percentage, 2
        ),

        status=status,

        total_questions=total_questions

    )


# ==========================
# RUN FLASK APPLICATION
# ==========================

if __name__ == "__main__":

    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True
    )


